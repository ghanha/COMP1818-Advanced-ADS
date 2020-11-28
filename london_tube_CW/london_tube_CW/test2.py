from collections import deque, namedtuple
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
from astropy.table import Table

inf = float('inf')
Edge = namedtuple('Edge', 'start, end, cost')

def make_edge(start, end, cost=1):
    return Edge(start, end, cost)

class Graph:
    def __init__(self, edges):
        wrong_edges = [i for i in edges if len(i) not in [2, 3]]
        if wrong_edges:
            raise ValueError('Wrong edges data: {}'.format(wrong_edges))

        self.edges = [make_edge(*edge) for edge in edges]

    @property
    def vertices(self):
        return set(
            sum(
                ([edge.start, edge.end] for edge in self.edges), []
            )
        )

    def get_node_pairs(self, n1, n2, both_ends=True):
        if both_ends:
            node_pairs = [[n1, n2], [n2, n1]]
        else:
            node_pairs = [[n1, n2]]
        return node_pairs

    def add_edge(self, n1, n2, cost=0, both_ends=True):
        node_pairs = self.get_node_pairs(n1, n2, both_ends)
        for edge in self.edges:
            if [edge.start, edge.end] in node_pairs:
                return ValueError('Edge {} {} already exists'.format(n1, n2))

        self.edges.append(Edge(start=n1, end=n2, cost=cost))
        if both_ends:
            self.edges.append(Edge(start=n2, end=n1, cost=cost))

    @property
    def neighbours(self):
        neighbours = {vertex: set() for vertex in self.vertices}
        for edge in self.edges:
            neighbours[edge.start].add((edge.end, edge.cost))

        return neighbours

    def dijkstra(self, source, dest):
        distances = {vertex: inf for vertex in self.vertices}
        previous_vertices = {
            vertex: None for vertex in self.vertices
        }
        distances[source] = 0
        vertices = self.vertices.copy()

        while vertices:
            current_vertex = min(vertices, key=lambda vertex: distances[vertex])
            vertices.remove(current_vertex)
            if distances[current_vertex] == inf:
                break
            for neighbour, cost in self.neighbours[current_vertex]:
                #c = cost.appe
                alternative_route = distances[current_vertex] + cost
                if alternative_route < distances[neighbour]:
                    distances[neighbour] = alternative_route
                    previous_vertices[neighbour] = current_vertex

        path, current_vertex = deque(), dest
        while previous_vertices[current_vertex] is not None:
            path.appendleft(current_vertex)
            current_vertex = previous_vertices[current_vertex]
        if path:
            path.appendleft(current_vertex)

        return distances[dest], path, #current_vertex.cost


wb = Workbook()
wb = load_workbook("London Underground data.xlsx")
ws = wb.active

book = openpyxl.load_workbook('London Underground data.xlsx')
sheet = book.active

stations = []
tuple_list = []

for row in sheet.iter_rows(min_row=1, min_col=2, max_row=754, max_col=4):
    stations_row = []
    tuple_row = ()

    for cell in row:
        stations_row.append(cell.value)
        tuple_row = tuple_row + (cell.value,)

    if stations_row[2] is not None:
        stations.append(stations_row)
        tuple_list.append(tuple_row)

        reversed_stations_row = stations_row[:]
        element0 = reversed_stations_row[0]
        reversed_stations_row[0] = reversed_stations_row[1]
        reversed_stations_row[1] = element0

        stations.append(reversed_stations_row)
        tuple_list.append(tuple(reversed_stations_row))

stations_final = Graph(stations)

from_station = input("From: ")
to_station = input("To: ")

final = stations_final.dijkstra(from_station,to_station)
print("Total time of your journey is:", final[0], "minutes")
print()
print("The number of the stations you will travel by is", len(final[1]),':')
#print(final)
final_list = list(final[1])
for i in final_list:
    print(i)

