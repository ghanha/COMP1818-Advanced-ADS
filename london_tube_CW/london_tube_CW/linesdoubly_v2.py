import doublelinklist
from tkinter import *
from openpyxl import Workbook
from openpyxl import load_workbook

root = Tk()
root.title('Journey Planner')
root.geometry('500x800')

wb = Workbook()
wb = load_workbook("London Underground data.xlsx")
ws = wb.active
#tube = load_workbook("London Underground data.xlsx")
#sheet = tube.active
column_a = ws['A']  #(Han) Get departure line_name
column_b = ws['B']  #(Han) Get arrival station
bakerloo = ws['A1']
central = ws['A50']
circle = ws['A148']
district = ws['A218']
hammersmith = ws['A337']
jubilee = ws['A394']
metropolitan = ws['A447']
northern = ws['A517']
piccadilly = ws['A618']
victoria = ws['A724']
waterloo = ws['A755']

class Line:

    class _line:
        def __init__(self, from_station, to_station, time):
            self.from_station = from_station
            self.to_station = to_station
            self.time = time

        def __str__(self):
            return str(self.from_station) + " - " + str(self.to_station)+" - " + str(self.time) + " min"

    def __init__(self):
        self.dblist = doublelinklist.DoublyLinkList()

    def add(self, from_station, to_station, time):
        if self.dblist.headval.datavalue == None:
            self.dblist.headval.datavalue = self._line(from_station, to_station, time)
        else:
            self.dblist.append(self._line(from_station, to_station, time))

    def show_all(self):
        self.dblist.listprint(self.dblist.headval)

def find_station(): #(Han) should be 'get_departure' function
    line_list = ''
    station = []
    station_list = Line()
    for cell in column_a:   #(Han) 'cell' currently is A1 - Barkerloo, I'll try drop-down menu later so users can choose another line
        #print(cell.value)
        if cell.value == bakerloo.value:
            line_list = f'{line_list + str(cell.value)}'
            label_a.config(text=line_list)
            print('\nThis is', line_list, 'line')
            for i in ws.iter_rows(min_row=26, max_row=49, min_col=2, max_col=4, values_only=True):
                station.append(i)
        if cell.value == central.value:
            line_list = f'{line_list + str(cell.value)}'
            print('\nThis is', line_list, 'line')
            for i in ws.iter_rows(min_row=99, max_row=147, min_col=2, max_col=4, values_only=True):
                station.append(i)
        else:
            print('Please check your line\' name again!')
        for el in station:
            station_list.add(el[0], el[1], el[2])
        station_list.show_all()
        break

'''vict_info()          #Hey Barbara, I'm so sorry for cut your Hammersmith and Victoria
# print(stations_v)     #Your idea was great and helpful for us
vict_full=Line()        #But I think instead of using 10x2 functions for 10 lines for both 'from' and 'to'
                        #We can use only get_departure and get_arrival functions
for el in stations_v:   #I'll text to the group, just in case, also leave comments here! love chu <3
    vict_full.add(el[0], el[1], el[2])

vict_full.show_all()'''

def get_departure():
    pass

ba = Button(root, text="From", command = find_station)  #(Han) Show time to travel between each 2 stops when you click the button
ba.pack(pady=20)
label_a = Label(root, text="")
label_a.pack(pady=20)

def get_arrival():
    line_list = ''
    for cell in column_b:
        #line_list = list(dict.fromkeys(line_list)) #Remove duplicates
        line_list = f'{line_list + str(cell.value)}\n'
        label_b.config(text=line_list)
ba = Button(root, text="To", command = get_arrival)
ba.pack(pady=20)
label_b = Label(root, text="")
label_b.pack(pady=20)

find_station()

root.mainloop()



