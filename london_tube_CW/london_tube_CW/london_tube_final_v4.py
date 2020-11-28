# Import libraries
from collections import defaultdict, namedtuple
from heapq import *
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
from astropy.table import Table # important to install astropy!!
from tkinter import *
from tkinter import ttk
import sys
import os



# Read data from excel file
wb = Workbook()
wb = load_workbook("London Underground data modified4.xlsx")
ws = wb.active

book = openpyxl.load_workbook('London Underground data modified4.xlsx')
sheet = book.active



# BE inputs
# from_station = input("From: ").title()
# to_station = input("To: ").title()



# Create nested list of all stations
stations = []

for row in sheet.iter_rows(min_row=1, min_col=2, max_row=754, max_col=4): # 757 last one, 49 bakerloo
    stations_row = []
    for cell in row:
        stations_row.append(cell.value)

    if stations_row[2] is not None:
        stations.append(stations_row)

        reversed_stations_row = stations_row[:]
        element0 = reversed_stations_row[0]
        reversed_stations_row[0] = reversed_stations_row[1]
        reversed_stations_row[1] = element0

        stations.append(reversed_stations_row)
#print("print stations", stations)



# Dijktra
def dijkstra(stationlist, fromstation, tostation):
    g = defaultdict(list)
    for l,r,c in stationlist:   #l-from_station, c-cost/time, r-to_station
        g[l].append((c,r))
    q, seen, mins = [(0,fromstation,())], set(), {fromstation: 0}
    while q:
        (cost,v1,path) = heappop(q)
        if v1 not in seen:
            seen.add(v1)
            path = (v1, path)
            if v1 == tostation: return (cost, path)

            for c, v2 in g.get(v1, ()):
                if v2 in seen: continue
                prev = mins.get(v2, None)
                next = cost + c
                if prev is None or next < prev:
                    mins[v2] = next
                    heappush(q, (next, v2, path))
        stations_list = list(mins.items())
    return float("inf"), stations_list[-1]



# Flatten nested list of journey stations
def flatten(object):
    for item in object:
        if isinstance(item, (list, tuple, set)):
            yield from flatten(item)
        else:
            yield item



# Create list of stations resulted from dijkstra
biglist = []
for element in stations:
    biglist.append(element)



# Tkinter inputs
root = Tk()
root.title('TFL Journey Planner')
root.geometry('800x600')

mainframe=Frame(root)
mainframe.pack(fill="both") #expand=True

label=Label(mainframe,text="Plan A Journey",bg="navy",fg="white",padx=5,pady=5)
label.config(font=("Arial",18))
label.pack(fill="x")

# Create entry
frStn_var = StringVar()
frStn_entry = Entry(mainframe, width=30, textvariable=frStn_var)
frStn_entry.pack(padx=10, pady=10)
frStn_var.set("From")

toStn_var = StringVar()
toStn_entry = Entry(mainframe, width=30, textvariable=toStn_var)
toStn_entry.pack(padx=10, pady=10)
toStn_var.set("To")

# Friendly disclaimer
reminder = Label(mainframe, width=30, text='Note: Station names are case-sensitive!', font='arial, 10', fg="gray")
reminder.pack()



# Create restart button
def restart_program():
    python = sys.executable
    os.execl(python, python, * sys.argv)



def findStn():
    # Get entry values
    from_station = frStn_var.get()      #.title()
    to_station = toStn_var.get()        #.title()
    stnLabel = Label(mainframe, text='Your journey is from ' + from_station + ' to ' + to_station, font='arial, 10')
    stnLabel.pack(padx=10)

    # Disable button
    jrneBtn.config(text='Your journey planned!', state=DISABLED)

    # Get journey result
    result = list(dijkstra(biglist, from_station, to_station))
    a = list(flatten(result))

    # Journey summary
    # result_stations
    result_stations = []
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=754, max_col=4):
        result_stations_row = []
        for cell in row:
            result_stations_row.append(cell.value)
        if (result_stations_row[1] in a and result_stations_row[2] in a and result_stations_row[3] is not None):
            result_stations.append(result_stations_row)
    print("result_stations:", result_stations)

    # Total time
    time_in_total = a[0]
    print("\nTotal time of your journey is:", time_in_total, "minutes")
    timeLabel = Label(mainframe, text='Total time of your journey is: ' + str(time_in_total) + ' minutes', font='arial, 10')
    timeLabel.pack(padx=10)

    # Number of stations
    stat_between = a[1:]
    print("\nThe number of the stations you will travel by:", len(stat_between)-1, "stations")
    numberStnLabel = Label(mainframe, text='The number of the stations you will travel by: ' + str(len(stat_between)-1) + ' stations', font='arial, 10')
    numberStnLabel.pack(padx=10)

    # BE departure - arrival stations
    stat_between_order = reversed(stat_between)
    print("\nStations between", from_station, "and", to_station, ":")
    fromtoLabel = Label(mainframe, text='Stops between ' + str(from_station) + ' and ' + str(to_station) + ': ' + '\n' + str(stat_between[::-1]), font='arial, 10')
    fromtoLabel.pack(padx=10)


    # BE station list
    for item in stat_between_order:
        print(item)

    # BE - create @time column for table
    c = result_stations
    minutes = []
    for i in c:
        minutes.append(i[3])

    # BE - create @totaltime column for table
    def sum_minutes(l):
        total = 0
        sum_min = []
        for val in l:
            total = total + val
            # print(total)
            sum_min.append(total)
        return sum_min
    list_min = sum_minutes(minutes)
    for i in range(len(c)):
        c[i].append(list_min[i])

    # BE - @table
    data_rows_new = c
    tt = Table(rows=data_rows_new, names=('Line:', 'From:', 'To:', 'Time between:', 'Time sum:'))
    print(tt)

    # FE - @table: Treeview
    jrneTable = ttk.Treeview(mainframe)
    jrneTable['column'] = ('Line', 'From', 'To', 'Time between', 'Time sum')

    jrneTable.column('#0', width=0, stretch=NO)
    jrneTable.column('Line', anchor=W, width=180)
    jrneTable.column('From', anchor=W, width=150)
    jrneTable.column('To', anchor=W, width=150)
    jrneTable.column('Time between', anchor=W, width=80)
    jrneTable.column('Time sum', anchor=W, width=80)

    jrneTable.heading('#0', text='', anchor=W)
    jrneTable.heading('Line', text='Line:', anchor=W)
    jrneTable.heading('From', text='From:', anchor=W)
    jrneTable.heading('To', text='To:', anchor=W)
    jrneTable.heading('Time between', text='Time between:', anchor=W)
    jrneTable.heading('Time sum', text='Time sum:', anchor=W)

    count = 0
    for record in c:
        jrneTable.insert(parent='', index='end', iid=count, text='', values=(record[0], record[1], record[2], record[3], record[4]))
        count += 1

    jrneTable.pack(pady=10)

    restartBtn = Button(mainframe, text="Restart", command=restart_program)
    restartBtn.pack()



# Create journey button
jrneBtn = Button(mainframe, text='Plan my journey', command=findStn)
jrneBtn.pack(padx=10, pady=5)



root.mainloop()




