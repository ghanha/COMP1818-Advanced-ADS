import doublelinklist
from openpyxl import Workbook
from openpyxl import load_workbook

class Line:

    class _line:
        def __init__(self, from_station, to_station, time):
            self.from_station = from_station
            self.to_station = to_station
            self.time = time
            # we can also add interchanges here self.interchanges

        def __str__(self):
            return str(self.from_station) + " - " + str(self.to_station)+" - " + str(self.time) + " min"

    def __init__(self):
        self.dblist = doublelinklist.DoublyLinkList()

    def add(self, from_station, to_station, time):
        if self.dblist.headval.datavalue == None:
            self.dblist.headval.datavalue = self._line(from_station, to_station, time)
        else:
            self.dblist.append(self._line(from_station, to_station, time))

    def show_all(self):
        self.dblist.listprint(self.dblist.headval)

wb = Workbook()
ws = wb.active
tube = load_workbook("London Underground data.xlsx")
sheet = tube.active

print()
print("VICTORIA LINE:")

stations_v=[]
def vict_info():
    for i in sheet.iter_rows(min_row=740, max_row=754, min_col=2, max_col=4,values_only=True):
        stations_v.append(i)
    return

vict_info()

#print(stations_v)

vict_full=Line()

for el in stations_v:
    vict_full.add(el[0], el[1], el[2])

vict_full.show_all()


print()
print("JUBILEE LINE:")
stations_j=[]
def jub_info():
    for i in sheet.iter_rows(min_row=421, max_row=446, min_col=2, max_col=4,values_only=True):
        stations_j.append(i)
    return

jub_info()
jub_full=Line()

for el in stations_j:
    jub_full.add(el[0], el[1], el[2])

jub_full.show_all()

print()
print("HAMMERSMITH & CITY LINE:")
stations_h=[]
def hamm_info():
    for i in sheet.iter_rows(min_row=366, max_row=393, min_col=2, max_col=4,values_only=True):
        stations_h.append(i)
    return

hamm_info()
hamm_full=Line()

for el in stations_h:
    hamm_full.add(el[0], el[1], el[2])

hamm_full.show_all()


